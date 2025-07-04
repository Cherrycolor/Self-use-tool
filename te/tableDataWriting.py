import os
import random
import pandas as pd
from openpyxl import load_workbook
from faker import Faker
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import numpy as np

"""
prd文档：linhuiba租户crm历史成交客户清洗
prd链接：https://confluence.lanhanba.com/pages/viewpage.action?pageId=149531956
"""

# 初始化Faker生成中文数据
fake = Faker('zh_CN')

# 硬编码所有下拉选项数据（从模板解析结果中提取）
BIRTHPLACE_OPTIONS = [
    '北京市','天津市','石家庄市','唐山市','秦皇岛市','邯郸市','邢台市','保定市','张家口市','承德市','沧州市',
    '廊坊市','衡水市','太原市','大同市','阳泉市','长治市','晋城市','朔州市','晋中市','运城市','忻州市','临汾市','吕梁市','呼和浩特市',
    '包头市','乌海市','赤峰市','通辽市','鄂尔多斯市','呼伦贝尔市','巴彦淖尔市','乌兰察布市','兴安盟','锡林郭勒盟','阿拉善盟','沈阳市',
    '大连市','鞍山市','抚顺市','本溪市','丹东市','锦州市','营口市','阜新市','辽阳市','盘锦市','铁岭市','朝阳市','葫芦岛市','长春市',
    '吉林市','四平市','辽源市','通化市','白山市','松原市','白城市','延边朝鲜族自治州','哈尔滨市','齐齐哈尔市','鸡西市','鹤岗市','双鸭山市',
    '大庆市','伊春市','佳木斯市','七台河市','牡丹江市','黑河市','绥化市','大兴安岭地区','上海市','南京市','无锡市','徐州市','常州市','苏州市',
    '南通市','连云港市','淮安市','盐城市','扬州市','镇江市','泰州市','宿迁市','杭州市','宁波市','温州市','嘉兴市','湖州市','绍兴市','金华市',
    '衢州市','舟山市','台州市','丽水市','合肥市','芜湖市','蚌埠市','淮南市','马鞍山市','淮北市','铜陵市','安庆市','黄山市','滁州市','阜阳市',
    '宿州市','六安市','亳州市','池州市','宣城市','福州市','厦门市','莆田市','三明市','泉州市','漳州市','南平市','龙岩市','宁德市','南昌市',
    '景德镇市','萍乡市','九江市','新余市','鹰潭市','赣州市','吉安市','宜春市','抚州市','上饶市','济南市','青岛市','淄博市','枣庄市','东营市',
    '烟台市','潍坊市','济宁市','泰安市','威海市','日照市','临沂市','德州市','聊城市','滨州市','菏泽市','郑州市','开封市','洛阳市','平顶山市',
    '安阳市','鹤壁市','新乡市','焦作市','濮阳市','许昌市','漯河市','三门峡市','南阳市','商丘市','信阳市','周口市','驻马店市','济源市','武汉市',
    '黄石市','十堰市','宜昌市','襄阳市','鄂州市','荆门市','孝感市','荆州市','黄冈市','咸宁市','随州市','恩施土家族苗族自治州','仙桃市','潜江市',
    '天门市','神农架林区','长沙市','株洲市','湘潭市','衡阳市','邵阳市','岳阳市','常德市','张家界市','益阳市','郴州市','永州市','怀化市','娄底市',
    '湘西土家族苗族自治州','广州市','韶关市','深圳市','珠海市','汕头市','佛山市','江门市','湛江市','茂名市','肇庆市','惠州市','梅州市','汕尾市',
    '河源市','阳江市','清远市','东莞市','中山市','潮州市','揭阳市','云浮市','南宁市','柳州市','桂林市','梧州市','北海市','防城港市','钦州市','贵港市',
    '玉林市','百色市','贺州市','河池市','来宾市','崇左市','海口市','三亚市','三沙市','儋州市','五指山市','琼海市','文昌市','万宁市','东方市','定安县',
    '屯昌县','澄迈县','临高县','白沙黎族自治县','昌江黎族自治县','乐东黎族自治县','陵水黎族自治县','保亭黎族苗族自治县','琼中黎族苗族自治县','重庆市',
    '重庆郊县','成都市','自贡市','攀枝花市','泸州市','德阳市','绵阳市','广元市','遂宁市','内江市','乐山市','南充市','眉山市','宜宾市','广安市','达州市',
    '雅安市','巴中市','资阳市','阿坝藏族羌族自治州','甘孜藏族自治州','凉山彝族自治州','贵阳市','六盘水市','遵义市','安顺市','毕节市','铜仁市','黔西南布依族苗族自治州',
    '黔东南苗族侗族自治州','黔南布依族苗族自治州','昆明市','曲靖市','玉溪市','保山市','昭通市','丽江市','普洱市','临沧市','楚雄彝族自治州','红河哈尼族彝族自治州','文山壮族苗族自治州',
    '西双版纳傣族自治州','大理白族自治州','德宏傣族景颇族自治州','怒江傈僳族自治州','迪庆藏族自治州','拉萨市','日喀则市','昌都市','林芝市','山南市','那曲市','阿里地区','西安市','铜川市',
    '宝鸡市','咸阳市','渭南市','延安市','汉中市','榆林市','安康市','商洛市','兰州市','嘉峪关市','金昌市','白银市','天水市','武威市','张掖市','平凉市','酒泉市','庆阳市','定西市','陇南市',
    '临夏回族自治州','甘南藏族自治州','西宁市','海东市','海北藏族自治州','黄南藏族自治州','海南藏族自治州','果洛藏族自治州','玉树藏族自治州','海西蒙古族藏族自治州','银川市','石嘴山市','吴忠市','固原市',
    '中卫市','乌鲁木齐市','克拉玛依市','吐鲁番市','哈密市','昌吉回族自治州','博尔塔拉蒙古自治州','巴音郭楞蒙古自治州','阿克苏地区','克孜勒苏柯尔克孜自治州','喀什地区','和田地区',
    '伊犁哈萨克自治州','塔城地区','阿勒泰地区','石河子市','阿拉尔市','图木舒克市','五家渠市','北屯市','铁门关市','双河市','可克达拉市','昆玉市','胡杨河市','香港特别行政区','台北市','澳门特别行政区'
]

INDUSTRY_OPTIONS = [
    ("汽车", "汽车销售"), ("汽车", "汽车租赁"), ("汽车", "汽车服务"),
    ("房地产", "房产销售"), ("房地产", "房产租赁"), ("房地产", "房产服务"),
    ("食品酒饮", "休闲食品"), ("食品酒饮", "粮油调味"), ("食品酒饮", "生鲜蔬菜"),
    ("食品酒饮", "烟酒饮料"), ("个护清洁", "个人护理"), ("个护清洁", "家庭护理"),
    ("电子产品", "手机"), ("电子产品", "电脑办公"), ("电子产品", "数码相机"),
    ("电子产品", "影音娱乐"), ("电子产品", "智能设备"), ("教育培训", "教育培训"),
    ("钟表珠宝", "钟表珠宝"), ("服饰鞋包", "男装女装"), ("服饰鞋包", "男鞋女鞋"),
    ("服饰鞋包", "箱包"), ("服饰鞋包", "服饰配件"), ("母婴亲子", "母婴用品"),
    ("母婴亲子", "母婴服务"), ("互联网", "基础服务"), ("互联网", "商务应用"),
    ("互联网", "交流娱乐"), ("互联网", "媒体"), ("互联网", "共享经济")
]

BUSINESS_OPTIONS = [
    ("零售", "女装"), ("零售", "男装"), ("零售", "男女装集合"), ("零售", "其他服饰"),
    ("零售", "鞋品箱包"), ("零售", "烟酒"), ("零售", "茶叶"), ("零售", "水果"),
    ("零售", "零食"), ("零售", "饮品"), ("零售", "米面粮油"), ("零售", "其他食品"),
    ("零售", "数码电子"), ("零售", "家电"), ("零售", "黄金珠宝"), ("零售", "钟表"),
    ("零售", "配饰"), ("零售", "杂品"), ("零售", "大型连锁超市"), ("零售", "便利店"),
    ("零售", "小型超市"), ("零售", "生鲜超市"), ("零售", "花店"), ("零售", "免税店"),
    ("零售", "OUTLETS"), ("零售", "家居家用"), ("零售", "商用设备"), ("零售", "农用设备"),
    ("零售", "工业设备"), ("零售", "五金建材"), ("零售", "办公耗材"), ("零售", "运动户外"),
    ("零售", "自行车"), ("零售", "电瓶车"), ("零售", "其他出行工具"), ("零售", "个护美妆"),
    ("零售", "药品器械"), ("零售", "保健品"), ("餐饮", "地方菜系"), ("餐饮", "异国餐饮"),
    ("餐饮", "自助餐"), ("餐饮", "烧烤"), ("餐饮", "锅类"), ("餐饮", "小吃快餐"),
    ("餐饮", "咖啡"), ("餐饮", "茶饮"), ("餐饮", "其他饮品"), ("餐饮", "烘焙"),
    ("餐饮", "甜品"), ("餐饮", "美食广场"), ("餐饮", "主题餐厅"), ("餐饮", "酒吧")
]


def generate_random_brand_data(num_records=100):
    """生成随机品牌数据"""
    brands = []

    for _ in range(num_records):
        # 随机选择行业和业态
        primary_industry, secondary_industry = random.choice(INDUSTRY_OPTIONS)
        primary_business, secondary_business = random.choice(BUSINESS_OPTIONS)

        # 随机生成成立时间 (1949-2023年)
        founding_year = random.randint(1949, 2023)
        founding_date = f"{founding_year}"

        brand = {
            "*品牌名称": f"{fake.company_prefix()}",
            "一级所属行业": primary_industry,
            "二级所属行业": secondary_industry,
            "*品牌类型": random.choice(["买家品牌", "供应商品牌","服务商品牌","其他品牌"]),
            "成立时间": founding_date,
            "发源地": random.choice(BIRTHPLACE_OPTIONS),
            "所属公司": f"{fake.company()}",
            "品牌logo（最多1张）": f"https://example.com/logos/{fake.uuid4()}.png",
            "品牌icon（最多1张）": f"https://example.com/icons/{fake.uuid4()}.svg",
            "门店照片（最多30张）": ", ".join([
                f"https://example.com/stores/{fake.uuid4()}.jpg"
                for _ in range(random.randint(3, 10))
            ]),
            "品牌简介": " ".join(fake.sentences(nb=3)),
            "一级业态": primary_business,
            "二级业态": secondary_business
        }
        brands.append(brand)

    return brands


def write_brand_data_to_excel(template_path, output_path, num_records=100):
    """将品牌数据写入Excel模板"""
    try:
        # 生成随机品牌数据
        brand_data = generate_random_brand_data(num_records)

        # 加载模板文件
        wb = load_workbook(template_path)

        # 获取Sheet1
        ws = wb['Sheet1']

        # 查找数据开始行（跳过标题行）
        start_row = 3

        # 清除现有数据（保留标题）
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, 14):  # 清除前13列
                ws.cell(row=row, column=col).value = None

        # 写入新数据
        for row_idx, brand in enumerate(brand_data, start=start_row):
            ws.cell(row=row_idx, column=1).value = brand["*品牌名称"]
            ws.cell(row=row_idx, column=2).value = brand["一级所属行业"]
            ws.cell(row=row_idx, column=3).value = brand["二级所属行业"]
            ws.cell(row=row_idx, column=4).value = brand["*品牌类型"]
            ws.cell(row=row_idx, column=5).value = brand["成立时间"]
            ws.cell(row=row_idx, column=6).value = brand["发源地"]
            ws.cell(row=row_idx, column=7).value = brand["所属公司"]
            ws.cell(row=row_idx, column=8).value = brand["品牌logo（最多1张）"]
            ws.cell(row=row_idx, column=9).value = brand["品牌icon（最多1张）"]
            ws.cell(row=row_idx, column=10).value = brand["门店照片（最多30张）"]
            ws.cell(row=row_idx, column=11).value = brand["品牌简介"]
            ws.cell(row=row_idx, column=12).value = brand["一级业态"]
            ws.cell(row=row_idx, column=13).value = brand["二级业态"]

            # 设置自动换行格式
            for col in [10, 11]:  # 门店照片和品牌简介列需要自动换行
                ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True)

        # 自动调整列宽
        for col in range(1, 14):
            column_letter = get_column_letter(col)
            max_length = 0

            # 获取列标题
            header = ws.cell(row=4, column=col).value
            if header:
                max_length = len(str(header))

            # 获取列内容最大长度
            for row in range(start_row, start_row + len(brand_data)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    # 对于多行文本内容，取最长行
                    lines = str(cell_value).split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)

            # 设置列宽（略微宽松）
            adjusted_width = max_length + 3
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

        # 保存到新文件
        wb.save(output_path)
        print(f"✅ 成功生成 {len(brand_data)} 条品牌数据")
        print(f"💾 文件已保存至: {output_path}")
        return True

    except Exception as e:
        print(f"❌ 处理文件时发生错误: {e}")
        print("⚠️ 尝试创建新模板并写入数据...")

        # 创建新的工作簿
        wb = load_workbook()

        # 创建Sheet1
        ws = wb.active
        ws.title = "Sheet1"

        # 添加标题行
        headers = [
            "*品牌名称", "一级所属行业", "二级所属行业", "*品牌类型", "成立时间",
            "发源地", "所属公司", "品牌logo（最多1张）", "品牌icon（最多1张）",
            "门店照片（最多30张）", "品牌简介", "一级业态", "二级业态"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=4, column=col_idx, value=header)

        # 生成随机品牌数据
        brand_data = generate_random_brand_data(num_records)

        # 写入数据
        for row_idx, brand in enumerate(brand_data, start=5):
            ws.cell(row=row_idx, column=1).value = brand["*品牌名称"]
            ws.cell(row=row_idx, column=2).value = brand["一级所属行业"]
            ws.cell(row=row_idx, column=3).value = brand["二级所属行业"]
            ws.cell(row=row_idx, column=4).value = brand["*品牌类型"]
            ws.cell(row=row_idx, column=5).value = brand["成立时间"]
            ws.cell(row=row_idx, column=6).value = brand["发源地"]
            ws.cell(row=row_idx, column=7).value = brand["所属公司"]
            ws.cell(row=row_idx, column=8).value = brand["品牌logo（最多1张）"]
            ws.cell(row=row_idx, column=9).value = brand["品牌icon（最多1张）"]
            ws.cell(row=row_idx, column=10).value = brand["门店照片（最多30张）"]
            ws.cell(row=row_idx, column=11).value = brand["品牌简介"]
            ws.cell(row=row_idx, column=12).value = brand["一级业态"]
            ws.cell(row=row_idx, column=13).value = brand["二级业态"]

        # 保存新文件
        wb.save(output_path)
        print(f"✅ 已创建新模板并写入 {len(brand_data)} 条品牌数据")
        return True


# 文件路径配置
# 获取当前日期和时间（包含年月日时分秒和微秒）
current_time = datetime.now() # 输出示例：2025-07-03 14:30:45.123456
# 格式化输出（自定义格式）
formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
template_path = '/Users/xaioyang/Downloads/品牌导入模版 (1).xlsx'
output_path = f'/Users/xaioyang/Downloads/{formatted_time}.xlsx'

# 生成随机数据并写入Excel
write_brand_data_to_excel(template_path, output_path, num_records=10)

print("\n操作完成！")