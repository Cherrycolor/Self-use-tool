import os
import random
import pandas as pd
from openpyxl import load_workbook
from faker import Faker
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def clear_brand_data_to_excel(template_path):
    """将品牌数据写入Excel模板"""
    try:
        # 加载模板文件
        wb = load_workbook(template_path)

        # 获取Sheet1
        ws = wb['Sheet1']

        # 查找数据开始行（跳过标题行）
        start_row = 3

        # 清除现有数据（保留标题）
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, 20):  # 清除前20列
                ws.cell(row=row, column=col).value = None

        # 保存
        wb.save(template_path)
        print(f"💾 文件已清理成功！！")
        return True

    except Exception as e:
        print(f"❌ 处理文件时发生错误: {e}")
        return None


# 文件路径配置
template_path = "/Users/xaioyang/Downloads/品牌导入异常数据.xlsx"
if input("是否需要输入模版路径 y/n:") == 'y':
    template_path = input("请输入模板路径：").strip('" ')

# 生成随机数据并写入Excel
clear_brand_data_to_excel(template_path)

print("\n操作完成！")